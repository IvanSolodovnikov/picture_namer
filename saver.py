import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


class ExcelResultsSaver:
    def __init__(self, output_file='results.xlsx'):
        self.output_file = output_file
        self.results = []

    def add_result(self, image_path, found_text):
        """Добавляет результат в память"""
        self.results.append({
            'image_path': image_path,
            'text': found_text
        })

    def save_all(self):
        """Сохраняет все результаты в Excel"""
        if not self.results:
            print("Нет результатов для сохранения")
            return False

        try:
            # Создаем новую книгу
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Результаты"

            # Заголовки
            headers = ["Имя файла", "Текст"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
                ws.column_dimensions[get_column_letter(col_num)].width = 30

            # Заполняем данные
            for row_num, result in enumerate(self.results, start=2):
                # Имя файла
                ws.cell(row=row_num, column=1, value=os.path.basename(result['image_path']))

                # Текст
                ws.cell(row=row_num, column=2, value=result['text'])
                ws.cell(row=row_num, column=2).alignment = Alignment(wrap_text=True)

                # Изображение
                if os.path.exists(result['image_path']):
                    try:
                        img = Image(result['image_path'])
                        # Масштабируем
                        img.width = 150
                        img.height = 150
                        ws.add_image(img, f'C{row_num}')
                        ws.row_dimensions[row_num].height = 120
                    except Exception as img_error:
                        ws.cell(row=row_num, column=3, value=f"Ошибка: {str(img_error)[:50]}")
                else:
                    ws.cell(row=row_num, column=3, value="Файл не найден")

            # Сохраняем
            wb.save(self.output_file)
            print(f"Все результаты сохранены в {self.output_file}")
            return True

        except PermissionError:
            print(f"Ошибка: Нет доступа к файлу {self.output_file}. Закройте файл если он открыт.")
        except Exception as e:
            print(f"Критическая ошибка при сохранении: {e}")
        return False
